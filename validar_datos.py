# -*- coding: utf-8 -*-
"""
ETAPA 1 - Validación del cruce de datos
Academia Horizonte - certificados por cohorte

Solo LECTURA: lee los Excel de INSUMOS/ y no modifica nada.
"""
from openpyxl import load_workbook

# ============================================================
# BODEGA: rutas y umbrales de las reglas de negocio (AGENTS.md)
# ============================================================
RUTA_MAESTRO = "INSUMOS/Maestro_Estudiantes.xlsx"   # lista oficial de estudiantes
RUTA_NOTAS   = "INSUMOS/Registro_Evaluaciones.xlsx" # notas y asistencia por modulo

UMBRAL_PROMEDIO   = 70  # nota minima para aprobar
UMBRAL_ASISTENCIA = 80  # asistencia minima para tener certificado


# ============================================================
# FUNCION: leer_maestro()
#   Lee el maestro y devuelve un diccionario:
#   identificacion (str) -> {nombre, programa, cohorte}
# ============================================================
def leer_maestro():
    libro = load_workbook(RUTA_MAESTRO, data_only=True)
    hoja = libro.active

    estudiantes = {}  # VARIABLE: diccionario que se va llenando

    # BUCLE: una vuelta por cada fila del maestro (la fila 1 es encabezado)
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        identificacion = str(fila[0])  # celda A: Identificacion (viene como texto)
        estudiantes[identificacion] = {
            "nombre": fila[1],   # celda B: Nombre_Completo
            "programa": fila[3], # celda D: Programa
            "cohorte": fila[4],  # celda E: Cohorte
        }

    return estudiantes


# ============================================================
# FUNCION: leer_evaluaciones()
#   Lee el registro y devuelve una lista de registros.
#   Cada registro es un diccionario con los datos de UNA fila:
#   una nota y una asistencia de un modulo cursado.
# ============================================================
def leer_evaluaciones():
    libro = load_workbook(RUTA_NOTAS, data_only=True)
    hoja = libro.active

    evaluaciones = []  # VARIABLE: lista vacia

    # BUCLE: una vuelta por cada fila del registro de evaluaciones
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        evaluaciones.append({
            "identificacion": str(fila[0]),  # Identificacion
            "programa": fila[1],             # Programa
            "modulo": fila[2],               # Modulo
            "nota": fila[3],                 # Nota (0-100)
            "asistencia": fila[4],           # Asistencia_Pct (0-100)
        })

    return evaluaciones


# ============================================================
# FUNCION: calcular_promedio(lista_valores)
#   Devuelve el promedio de una lista de numeros.
#   Promedio = suma de los valores / cantidad de valores.
# ============================================================
def calcular_promedio(lista_valores):
    suma = 0  # VARIABLE: acumulador

    # BUCLE: suma uno por uno todos los valores de la lista
    for valor in lista_valores:
        suma += valor

    # DIVISION: el promedio es la suma dividida entre la cantidad
    return suma / len(lista_valores)


# ============================================================
# FUNCION: decidir_certificado(promedio, asistencia)
#   Aplica las reglas de AGENTS.md y devuelve el tipo:
#   "Aprobado", "Participación" o "Sin certificado".
#   Los limites INCLUYEN el valor (>= 70, >= 80).
# ============================================================
def decidir_certificado(promedio, asistencia):
    # CONDICIONAL 1: aprueba si el promedio alcanza 70 Y la asistencia 80
    if promedio >= UMBRAL_PROMEDIO and asistencia >= UMBRAL_ASISTENCIA:
        return "Aprobado"

    # CONDICIONAL 2: con la asistencia en 80 o mas (aunque el promedio
    # quede bajo 70) se otorga certificado de participacion
    if asistencia >= UMBRAL_ASISTENCIA:
        return "Participación"

    # CONDICIONAL 3: si la asistencia no llega a 80, no hay certificado
    return "Sin certificado"


# ============================================================
# FUNCION: imprimir_tabla(resultados)
#   Muestra la tabla de resultados ordenada por programa y nombre.
#   Solo muestra datos; no calcula nada.
# ============================================================
def imprimir_tabla(resultados):
    # Encabezado de la tabla con columnas fijas
    print("Identificación  Nombre                     Programa                       Módulos  Promedio  Asistencia  Certificado")
    print("=" * 120)

    # BUCLE: una fila por resultado, en el orden ya ordenado
    for r in resultados:
        # VARIABLES: valores formateados para la columna
        promedio_txt = f"{r['promedio']:.1f}".replace(".", ",")     # 83,8
        asistencia_txt = f"{r['asistencia']:.1f}".replace(".", ",") # 90,0

        # CONDICIONAL: el nombre se muestra "sin datos" si no hay ficha
        nombre = r["nombre"] if r["nombre"] else "(sin nombre en maestro)"

        print(
            f"{r['identificacion']:<14} "
            f"{nombre:<25} "
            f"{r['programa']:<27} "
            f"{r['modulos']:<8} "
            f"{promedio_txt:<9} "
            f"{asistencia_txt:<11} "
            f"{r['certificado']}"
        )


# ============================================================
# PROGRAMA PRINCIPAL
# ============================================================

# --- PASO 1 y 2: lectura y resumen de la bodega ---
maestro = leer_maestro()          # FUNCION: maestro
evaluaciones = leer_evaluaciones()  # FUNCION: evaluaciones

print("=" * 120)
print("1) RESUMEN DE LOS DATOS DE ENTRADA")
print("=" * 120)

# VARIABLES: contadores del resumen
total_maestro = len(maestro)
total_evaluaciones = len(evaluaciones)

# VARIABLE: diccionario para contar estudiantes por programa del maestro
por_programa = {}

# BUCLE: cuenta los estudiantes de cada programa del maestro
for datos in maestro.values():
    programa = datos["programa"]
    # CONDICIONAL: si el programa no se habia contado, se crea con 0
    if programa not in por_programa:
        por_programa[programa] = 0
    por_programa[programa] += 1

print(f"Maestro_Estudiantes.xlsx : {total_maestro} estudiantes (cohorte 2026-A)")
print("Estudiantes por programa:")
for programa, cantidad in por_programa.items():
    print(f"  - {programa}: {cantidad}")

print(f"Registro_Evaluaciones.xlsx: {total_evaluaciones} evaluaciones (filas)")

# VARIABLES: colecciones para los resumenes de evaluaciones
modulos_por_programa = {}  # programa -> conjunto de modulos vistos
notas = []                 # lista con todas las notas
asistencias = []           # lista con todas las asistencias

# BUCLE: recorre todas las evaluaciones para los tres resumenes
for ev in evaluaciones:
    programa = ev["programa"]
    # CONDICIONAL: crea el conjunto de modulos si el programa es nuevo
    if programa not in modulos_por_programa:
        modulos_por_programa[programa] = set()
    modulos_por_programa[programa].add(ev["modulo"])
    notas.append(ev["nota"])
    asistencias.append(ev["asistencia"])

# BUCLE: muestra los modulos que existen en cada programa
for programa, modulos in modulos_por_programa.items():
    modulos_ordenados = sorted(modulos)  # VARIABLE: modulos en orden
    print(f"  Modulos de {programa}: {', '.join(modulos_ordenados)}")

# VARIABLES: rangos minimo-maximo de notas y asistencias
rango_notas = f"{min(notas)}-{max(notas)}"
rango_asistencia = f"{min(asistencias)}-{max(asistencias)}"
print(f"Rango de Notas: {rango_notas} | Rango de Asistencia_Pct: {rango_asistencia}")

# --- PASO 3: cruce por Identificacion + Programa ---

# VARIABLE: diccionario para agrupar las evaluaciones
# clave = (identificacion, programa) -> lista de registros
grupos = {}

# BUCLE: reparte cada evaluacion en su grupo
for ev in evaluaciones:
    clave = (ev["identificacion"], ev["programa"])
    # CONDICIONAL: si el grupo no existe todavia, se crea
    if clave not in grupos:
        grupos[clave] = []
    grupos[clave].append(ev)

# --- PASO 4: calculo por estudiante con las funciones pedidas ---

resultados = []  # VARIABLE: lista final de resultados

# BUCLE: un resultado por cada grupo (estudiante + programa)
for (identificacion, programa), registros in grupos.items():
    # VARIABLE: datos del estudiante (None si no hay ficha en el maestro)
    datos_estudiante = maestro.get(identificacion)

    # VARIABLES: listas con las notas y asistencias del grupo
    notas_grupo = [r["nota"] for r in registros]
    asistencias_grupo = [r["asistencia"] for r in registros]

    # FUNCIONES: los dos promedios que piden las reglas de negocio
    promedio = calcular_promedio(notas_grupo)
    asistencia_promedio = calcular_promedio(asistencias_grupo)

    # FUNCION: tipo de certificado segun las reglas de AGENTS.md
    certificado = decidir_certificado(promedio, asistencia_promedio)

    # Se guarda el resultado con todo lo que pide la tabla
    resultados.append({
        "identificacion": identificacion,
        "nombre": datos_estudiante["nombre"] if datos_estudiante else None,
        "programa": programa,
        "modulos": len(registros),  # VARIABLE: cantidad de modulos cursados
        "promedio": promedio,
        "asistencia": asistencia_promedio,
        "certificado": certificado,
    })

# --- PASO 5: tabla ordenada por programa y nombre ---

# VARIABLE: copia ordenada de los resultados
# key = (programa, nombre o cadena vacia para los sin ficha)
ordenados = sorted(
    resultados,
    key=lambda r: (r["programa"], r["nombre"] or ""),
)

print()
print("=" * 120)
print("2) TABLA DE RESULTADOS (ordenada por programa y nombre)")
print("=" * 120)
imprimir_tabla(ordenados)  # FUNCION: imprime la tabla

# VARIABLES: contadores para el resumen final
conteos = {}

# BUCLE: cuenta cuantos hay de cada tipo de certificado
for r in resultados:
    certificado = r["certificado"]
    # CONDICIONAL: primer conteo de este tipo o suma uno
    if certificado not in conteos:
        conteos[certificado] = 0
    conteos[certificado] += 1

# VARIABLE: total de certificados que se emitirian
emitibles = conteos.get("Aprobado", 0) + conteos.get("Participación", 0)

print()
print(f"Resumen: {conteos.get('Aprobado', 0)} Aprobación | "
      f"{conteos.get('Participación', 0)} Participación | "
      f"{conteos.get('Sin certificado', 0)} Sin certificado "
      f"-> {emitibles} certificados emitibles")

# --- PASO 6: inconsistencias del cruce ---

# VARIABLES: conjuntos de identificaciones de cada fuente
ids_maestro = set(maestro.keys())
ids_con_notas = set(g for g, _ in grupos.keys())

# VARIABLES: las dos listas de inconsistencias
en_maestro_sin_notas = ids_maestro - ids_con_notas
con_notas_sin_maestro = ids_con_notas - ids_maestro

print()
print("=" * 120)
print("3) INCONSISTENCIAS EN EL CRUCE")
print("=" * 120)

print("a) Estudiantes del maestro SIN ninguna evaluacion:")
# CONDICIONAL: solo hay lista si existen casos
if en_maestro_sin_notas:
    # BUCLE: un caso por linea, con el nombre desde el maestro
    for identificacion in sorted(en_maestro_sin_notas):
        datos = maestro[identificacion]
        print(f"   - {identificacion} | {datos['nombre']} | {datos['programa']}")
else:
    print("   (ninguno)")

print("b) Identificaciones en evaluaciones que NO estan en el maestro:")
# CONDICIONAL: solo hay lista si existen casos
if con_notas_sin_maestro:
    # BUCLE: un caso por linea, con su programa desde las evaluaciones
    for identificacion in sorted(con_notas_sin_maestro):
        # VARIABLE: primer registro de ese id para saber el programa
        programa_ev = next(g[1] for g in grupos if g[0] == identificacion)
        print(f"   - {identificacion} | (sin nombre en maestro) | {programa_ev}")
else:
    print("   (ninguno)")
