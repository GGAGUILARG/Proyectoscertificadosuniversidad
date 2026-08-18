# -*- coding: utf-8 -*-
"""
APP: Academia Horizonte - Sistema de Certificados
=================================================

Tres piezas del proyecto:
  - Bodega   (datos)     : los Excel en la carpeta INSUMOS (no se modifican).
  - Cocina   (logica)    : este archivo, app.py. Todo el calculo vive aqui.
  - Salon    (interfaz)  : los HTML en templates/, que solo muestran datos.

Reglas de negocio (todos los limites INCLUYEN el valor):
  - Promedio   = suma de Notas / cantidad de modulos cursados.
  - Asistencia = promedio de Asistencia_Pct.
  - Aprobacion     si Promedio >= 70  y Asistencia >= 80.
  - Participacion  si Promedio < 70   y Asistencia >= 80.
  - Sin certificado si Asistencia < 80.
  - Un certificado por cada combinacion (Identificacion + Programa).
"""

import json
import os
from functools import wraps

from flask import Flask, render_template, abort, request, redirect, url_for, session, flash
from openpyxl import load_workbook
from pathlib import Path
from werkzeug.security import check_password_hash, generate_password_hash

# ============================================================
# BODEGA: rutas y datos de entrada (no se escriben, solo leen)
# ============================================================

# VARIABLE: carpeta donde vive este archivo (app.py), sin importar
# desde donde se ejecute la aplicacion (local, Vercel, otro servidor)
RAIZ = Path(__file__).resolve().parent

# Las rutas a los Excel se arman SIEMPRE a partir de RAIZ, asi la app
# encuentra la bodega aunque se ejecute desde otro directorio
RUTA_MAESTRO = RAIZ / "INSUMOS" / "Maestro_Estudiantes.xlsx"   # lista oficial de estudiantes
RUTA_NOTAS   = RAIZ / "INSUMOS" / "Registro_Evaluaciones.xlsx" # notas y asistencia por modulo

UMBRAL_PROMEDIO   = 70   # nota minima para aprobar
UMBRAL_ASISTENCIA = 80   # asistencia minima para tener certificado


def cargar_variables_locales_entorno():
    """
    Carga variables desde .env solo en desarrollo local.
    En Vercel las variables se definen en Project Settings.
    """
    ruta_env = RAIZ / ".env"
    if not ruta_env.exists():
        return

    try:
        lineas = ruta_env.read_text(encoding="utf-8").splitlines()
    except OSError:
        return

    for linea in lineas:
        texto = linea.strip()
        if not texto or texto.startswith("#"):
            continue

        if "=" not in texto:
            continue

        clave, valor = texto.split("=", 1)
        clave = clave.strip()
        valor = valor.strip()

        if not clave:
            continue

        os.environ.setdefault(clave, valor)


cargar_variables_locales_entorno()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "cambiar-esta-clave-en-entorno")

RUTA_USUARIOS = Path(os.getenv("RUTA_USUARIOS", "usuarios_locales.json"))
if not RUTA_USUARIOS.is_absolute():
    RUTA_USUARIOS = RAIZ / RUTA_USUARIOS


def normalizar_usuario(item):
    if not isinstance(item, dict):
        return None

    correo = str(item.get("correo", "")).strip().lower()
    rol = str(item.get("rol", "normal")).strip().lower()
    clave_hash = str(item.get("clave_hash", "")).strip()
    clave = str(item.get("clave", "")).strip()

    if not correo:
        return None

    if rol not in {"admin", "normal"}:
        rol = "normal"

    return {
        "correo": correo,
        "rol": rol,
        "clave_hash": clave_hash,
        "clave": clave,
    }


def leer_usuarios_desde_texto_json(texto):
    if not texto:
        return []

    try:
        datos = json.loads(texto)
    except json.JSONDecodeError:
        return []

    if not isinstance(datos, list):
        return []

    usuarios = []
    for item in datos:
        usuario = normalizar_usuario(item)
        if usuario is not None:
            usuarios.append(usuario)

    return usuarios


def obtener_admin_inicial_desde_entorno():
    correo = os.getenv("ADMIN_CORREO", "").strip().lower()
    clave_hash = os.getenv("ADMIN_CLAVE_HASH", "").strip()
    clave = os.getenv("ADMIN_CLAVE", "").strip()

    if not correo:
        return None

    if not clave_hash and not clave:
        return None

    return {
        "correo": correo,
        "rol": "admin",
        "clave_hash": clave_hash,
        "clave": clave,
    }


def guardar_usuarios_autorizados(usuarios):
    try:
        RUTA_USUARIOS.parent.mkdir(parents=True, exist_ok=True)
        with RUTA_USUARIOS.open("w", encoding="utf-8") as archivo:
            json.dump(usuarios, archivo, ensure_ascii=False, indent=2)
    except OSError:
        return False

    return True


def cargar_usuarios_autorizados():
    """
    Lee usuarios desde usuarios_locales.json (si existe).
    Si no existe, usa la variable de entorno USUARIOS_JSON.

    El admin inicial tambien puede llegar por entorno:
      ADMIN_CORREO + ADMIN_CLAVE_HASH (o ADMIN_CLAVE para pruebas).

    Formato esperado (lista JSON):
      [
        {"correo": "admin@academia.cr", "rol": "admin", "clave_hash": "..."},
        {"correo": "usuario@academia.cr", "rol": "normal", "clave_hash": "..."}
      ]

    Tambien acepta "clave" (texto plano) para pruebas locales,
    pero en produccion se recomienda usar solo "clave_hash".
    """
    if RUTA_USUARIOS.exists():
        try:
            texto_archivo = RUTA_USUARIOS.read_text(encoding="utf-8")
        except OSError:
            texto_archivo = ""
        usuarios = leer_usuarios_desde_texto_json(texto_archivo)
    else:
        texto_entorno = os.getenv("USUARIOS_JSON", "").strip()
        usuarios = leer_usuarios_desde_texto_json(texto_entorno)

    admin_inicial = obtener_admin_inicial_desde_entorno()
    if admin_inicial is not None:
        correos = {u["correo"] for u in usuarios}
        if admin_inicial["correo"] not in correos:
            usuarios.append(admin_inicial)

    return usuarios


def crear_usuario_normal(correo, clave):
    correo_limpio = correo.strip().lower()
    clave_limpia = clave.strip()

    if not correo_limpio or "@" not in correo_limpio:
        return False, "El correo no es válido."

    if len(clave_limpia) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres."

    usuarios = cargar_usuarios_autorizados()
    existe = any(u["correo"] == correo_limpio for u in usuarios)
    if existe:
        return False, "Ese correo ya tiene acceso."

    usuarios.append({
        "correo": correo_limpio,
        "rol": "normal",
        "clave_hash": generate_password_hash(clave_limpia),
        "clave": "",
    })

    if not guardar_usuarios_autorizados(usuarios):
        return False, "No se pudo guardar el usuario. En Vercel necesitás un almacenamiento externo para persistir altas."

    return True, "Acceso creado correctamente."


def listar_usuarios_publicos():
    usuarios = cargar_usuarios_autorizados()
    ordenados = sorted(usuarios, key=lambda u: (u["rol"], u["correo"]))
    return [{"correo": u["correo"], "rol": u["rol"]} for u in ordenados]


def validar_credenciales(correo, clave):
    correo_limpio = correo.strip().lower()

    for usuario in cargar_usuarios_autorizados():
        if usuario["correo"] != correo_limpio:
            continue

        if usuario["clave_hash"]:
            if check_password_hash(usuario["clave_hash"], clave):
                return {"correo": usuario["correo"], "rol": usuario["rol"]}
            return None

        if usuario["clave"] and usuario["clave"] == clave:
            return {"correo": usuario["correo"], "rol": usuario["rol"]}

        return None

    return None


def obtener_usuario_actual():
    correo = session.get("usuario_correo")
    rol = session.get("usuario_rol")

    if not correo or not rol:
        return None

    return {
        "correo": correo,
        "rol": rol,
        "es_admin": rol == "admin",
    }


def login_requerido(funcion_vista):
    @wraps(funcion_vista)
    def envoltura(*args, **kwargs):
        usuario = obtener_usuario_actual()
        if usuario is None:
            return redirect(url_for("pagina_login", siguiente=request.path))
        return funcion_vista(*args, **kwargs)

    return envoltura


def admin_requerido(funcion_vista):
    @wraps(funcion_vista)
    def envoltura(*args, **kwargs):
        usuario = obtener_usuario_actual()
        if usuario is None:
            return redirect(url_for("pagina_login", siguiente=request.path))

        if not usuario["es_admin"]:
            abort(403)

        return funcion_vista(*args, **kwargs)

    return envoltura


# ============================================================
# FUNCION: cargar_estudiantes()
#   Lee el maestro y devuelve un diccionario:
#   identificacion (str) -> {nombre, programa, cohorte}
#   El diccionario permite buscar a un estudiante de forma directa.
# ============================================================
def cargar_estudiantes():
    libro = load_workbook(RUTA_MAESTRO, data_only=True)
    hoja = libro.active

    estudiantes = {}  # VARIABLE: diccionario que se va llenando

    # BUCLE: una vuelta por cada fila del maestro (la fila 1 es encabezado)
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        identificacion = str(fila[0])  # celda A: Identificacion (viene como texto)
        estudiantes[identificacion] = {
            "nombre": fila[1],   # celda B: Nombre_Completo
            "programa": fila[3], # celda D: Programa
            "cohorte": fila[4],  # celda E: Cohorte
        }

    return estudiantes


# ============================================================
# FUNCION: cargar_notas()
#   Lee el registro de evaluaciones y devuelve una lista.
#   Cada elemento es un diccionario con los datos de UNA fila:
#   una nota y una asistencia de un modulo cursado.
# ============================================================
def cargar_notas():
    libro = load_workbook(RUTA_NOTAS, data_only=True)
    hoja = libro.active

    notas = []  # VARIABLE: lista vacia

    # BUCLE: una vuelta por cada fila del registro de evaluaciones
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        notas.append({
            "identificacion": str(fila[0]),  # Identificacion
            "programa": fila[1],             # Programa
            "modulo": fila[2],               # Modulo
            "nota": fila[3],                 # Nota (0-100)
            "asistencia": fila[4],           # Asistencia_Pct (0-100)
        })

    return notas


# ============================================================
# FUNCION: calcular_promedio(lista_valores)
#   Devuelve el promedio de una lista de numeros.
#   Promedio = suma de los valores / cantidad de valores.
# ============================================================
def calcular_promedio(lista_valores):
    suma = 0  # VARIABLE: acumulador

    # BUCLE: suma uno por uno todos los valores de la lista
    for valor in lista_valores:
        suma += valor

    # DIVISION: el promedio es la suma dividida entre la cantidad
    return suma / len(lista_valores)


# ============================================================
# FUNCION: decidir_certificado(promedio, asistencia)
#   Aplica las reglas de negocio y devuelve el estado:
#   "Aprobado", "Participacion" o "Sin certificado".
#   Los limites INCLUYEN el valor (>= 70, >= 80).
# ============================================================
def decidir_certificado(promedio, asistencia):
    # CONDICIONAL 1: aprueba si el promedio alcanza 70 Y la asistencia 80
    if promedio >= UMBRAL_PROMEDIO and asistencia >= UMBRAL_ASISTENCIA:
        return "Aprobado"

    # CONDICIONAL 2: con la asistencia en 80 o mas (aunque el promedio
    # quede bajo 70) se otorga certificado de participacion
    if asistencia >= UMBRAL_ASISTENCIA:
        return "Participacion"

    # CONDICIONAL 3: si la asistencia no llega a 80, no hay certificado
    return "Sin certificado"


# ============================================================
# FUNCION: calcular_resultados()
#   Cruza las dos fuentes y calcula un resultado por cada
#   estudiante + programa. Devuelve una lista ordenada por
#   programa y nombre.
#
#   Decision vigente (AGENTS.md): los estudiantes con notas pero
#   sin ficha en el maestro ("999880777") NO reciben certificado;
#   quedan fuera de los resultados, sin advertencias.
# ============================================================
def calcular_resultados():
    estudiantes = cargar_estudiantes()  # FUNCION: maestro
    notas = cargar_notas()              # FUNCION: notas

    # VARIABLE: diccionario para agrupar las notas
    # clave = (identificacion, programa) -> lista de registros
    grupos = {}

    # BUCLE: reparte cada nota en su grupo
    for registro in notas:
        clave = (registro["identificacion"], registro["programa"])

        # CONDICIONAL: si la clave no existe todavia, se crea
        if clave not in grupos:
            grupos[clave] = []

        grupos[clave].append(registro)

    resultados = []  # VARIABLE: lista final con los resultados

    # BUCLE: un resultado por cada grupo (estudiante + programa)
    for (identificacion, programa), registros_grupo in grupos.items():
        # VARIABLE: ficha del estudiante (None si no esta en el maestro)
        ficha = estudiantes.get(identificacion)

        # CONDICIONAL: sin ficha en el maestro no se emite certificado
        if ficha is None:
            continue  # salta este grupo y sigue con el siguiente

        # VARIABLES: listas con las notas y asistencias del grupo
        notas_grupo = [r["nota"] for r in registros_grupo]
        asistencias_grupo = [r["asistencia"] for r in registros_grupo]

        # FUNCIONES: los dos promedios que piden las reglas de negocio
        promedio = calcular_promedio(notas_grupo)
        asistencia = calcular_promedio(asistencias_grupo)

        # FUNCION: estado segun las reglas de negocio
        estado = decidir_certificado(promedio, asistencia)

        # Se arma el resultado completo para la interfaz
        resultados.append({
            "identificacion": identificacion,
            "nombre": ficha["nombre"],
            "programa": programa,
            "cohorte": ficha["cohorte"],
            "promedio": promedio,
            "asistencia": asistencia,
            "estado": estado,
            "modulos": registros_grupo,  # lista con modulo, nota y asistencia
        })

    # VARIABLE: copia ordenada por programa y luego por nombre
    ordenados = sorted(resultados, key=lambda r: (r["programa"], r["nombre"]))

    return ordenados


# ============================================================
# FUNCION: formatear_numero(valor)
#   Convierte un numero como 91.25 en texto con coma decimal
#   (estilo espanol) y dos decimales: "91,25".
# ============================================================
def formatear_numero(valor):
    return f"{valor:.2f}".replace(".", ",")


# ============================================================
# FUNCION: obtener_programas()
#   Devuelve la lista ordenada de programas del maestro.
# ============================================================
def obtener_programas():
    estudiantes = cargar_estudiantes()

    # VARIABLE: conjunto para no repetir programas
    programas = set()

    # BUCLE: recolecta el programa de cada estudiante
    for ficha in estudiantes.values():
        programas.add(ficha["programa"])

    # VARIABLE: lista ordenada alfabeticamente
    return sorted(programas)


# ============================================================
# FUNCION: calcular_inconsistencias()
#   Detecta los datos que no encajan en el cruce. Devuelve un
#   diccionario con tres listas (las 3 detectadas):
#     1. sin_evaluaciones: en el maestro, sin ninguna nota
#     2. sin_maestro     : con notas, pero sin ficha en el maestro
#     3. sin_certificado : asistencia final < 80 (no recibe certificado)
# ============================================================
def calcular_inconsistencias():
    estudiantes = cargar_estudiantes()
    notas = cargar_notas()
    resultados = calcular_resultados()

    # VARIABLES: conjunto de identificaciones de cada fuente
    ids_maestro = set(estudiantes)
    ids_con_notas = set(n["identificacion"] for n in notas)

    # VARIABLE: lista de estudiantes del maestro sin ninguna evaluacion
    lista_sin_evaluaciones = []

    # BUCLE: los ids que solo estan en el maestro
    for identificacion in sorted(ids_maestro - ids_con_notas):
        ficha = estudiantes[identificacion]
        lista_sin_evaluaciones.append({
            "identificacion": identificacion,
            "nombre": ficha["nombre"],
            "programa": ficha["programa"],
        })

    # VARIABLE: lista de identificaciones con notas pero sin ficha
    lista_sin_maestro = []

    # BUCLE: los ids que solo estan en las notas
    for identificacion in sorted(ids_con_notas - ids_maestro):
        # VARIABLE: programa y modulos de ese id, desde las notas
        programa = next(n["programa"] for n in notas if n["identificacion"] == identificacion)
        modulos = sum(1 for n in notas if n["identificacion"] == identificacion)
        lista_sin_maestro.append({
            "identificacion": identificacion,
            "programa": programa,
            "modulos": modulos,
        })

    # VARIABLE: lista de estudiantes que quedan sin certificado
    lista_sin_certificado = []

    # BUCLE: los resultados con estado "Sin certificado"
    for resultado in resultados:
        # CONDICIONAL: solo interesan los que no reciben certificado
        if resultado["estado"] == "Sin certificado":
            ultimo = resultado["modulos"][-1]  # VARIABLE: ultimo modulo cursado
            lista_sin_certificado.append({
                "identificacion": resultado["identificacion"],
                "nombre": resultado["nombre"],
                "promedio": formatear_numero(resultado["promedio"]),
                "asistencia": formatear_numero(resultado["asistencia"]),
                "ultima_nota": ultimo["nota"],
                "ultima_asistencia": ultimo["asistencia"],
                "ultimo_modulo": ultimo["modulo"],
            })

    # VARIABLE: diccionario final con las tres listas
    return {
        "sin_evaluaciones": lista_sin_evaluaciones,
        "sin_maestro": lista_sin_maestro,
        "sin_certificado": lista_sin_certificado,
    }


# ============================================================
# SALON: rutas web que muestran la informacion
# ============================================================


@app.route("/login", methods=["GET", "POST"])
def pagina_login():
    if obtener_usuario_actual() is not None:
        return redirect(url_for("pagina_listado"))

    if not cargar_usuarios_autorizados():
        flash("No hay usuarios configurados. Definí un admin con ADMIN_CORREO y ADMIN_CLAVE_HASH.")

    if request.method == "POST":
        correo = request.form.get("correo", "").strip()
        clave = request.form.get("clave", "")
        usuario = validar_credenciales(correo, clave)

        if usuario is None:
            flash("Correo o contraseña incorrectos.")
            return render_template("login.html", correo_intentado=correo)

        session["usuario_correo"] = usuario["correo"]
        session["usuario_rol"] = usuario["rol"]

        siguiente = request.args.get("siguiente")
        if siguiente and siguiente.startswith("/"):
            return redirect(siguiente)

        return redirect(url_for("pagina_listado"))

    return render_template("login.html", correo_intentado="")


@app.route("/logout", methods=["POST"])
@login_requerido
def cerrar_sesion():
    session.clear()
    return redirect(url_for("pagina_login"))


@app.route("/admin/usuarios", methods=["GET", "POST"])
@login_requerido
@admin_requerido
def pagina_admin_usuarios():
    usuario = obtener_usuario_actual()

    if request.method == "POST":
        correo = request.form.get("correo", "")
        clave = request.form.get("clave", "")

        creado, mensaje = crear_usuario_normal(correo, clave)
        flash(mensaje)

        if creado:
            return redirect(url_for("pagina_admin_usuarios"))

        usuarios = listar_usuarios_publicos()
        return render_template(
            "admin_usuarios.html",
            usuario=usuario,
            usuarios=usuarios,
            correo_intentado=correo,
        )

    usuarios = listar_usuarios_publicos()
    return render_template(
        "admin_usuarios.html",
        usuario=usuario,
        usuarios=usuarios,
        correo_intentado="",
    )

# RUTA: pagina principal con el listado de resultados
@app.route("/")
@login_requerido
def pagina_listado():
    usuario = obtener_usuario_actual()

    resultados = calcular_resultados()          # FUNCION: resultados validados
    es_admin = usuario["es_admin"]

    if es_admin:
        inconsistencias = calcular_inconsistencias() # FUNCION: datos que no encajan
    else:
        inconsistencias = {
            "sin_evaluaciones": [],
            "sin_maestro": [],
            "sin_certificado": [],
        }

    programas = obtener_programas()              # FUNCION: programas del maestro

    # VARIABLE: filtro elegido en la pagina (?programa=...), "todos" por defecto
    programa_filtro = request.args.get("programa", "todos")

    # CONDICIONAL: segun el filtro se arma la tabla a mostrar
    if programa_filtro == "todos":
        tabla = resultados
    else:
        # VARIABLE: lista con solo los resultados del programa elegido
        tabla = [r for r in resultados if r["programa"] == programa_filtro]

    # VARIABLES: contadores del resumen (siempre del total, sin filtro)
    aprobados = 0
    participaciones = 0
    sin_certificado = 0

    # BUCLE: cuenta cuantos hay de cada estado
    for resultado in resultados:
        # CONDICIONAL: segun el estado se suma a un contador
        if resultado["estado"] == "Aprobado":
            aprobados += 1
        elif resultado["estado"] == "Participacion":
            participaciones += 1
        else:
            sin_certificado += 1

    resumen = {
        "aprobados": aprobados,
        "participaciones": participaciones,
        "sin_certificado": sin_certificado,
        "total": len(resultados),
    }

    # El HTML (salon) solo recibe datos ya calculados; no calcula nada
    return render_template(
        "index.html",
        resultados=tabla,
        resumen=resumen,
        inconsistencias=inconsistencias,
        programas=programas,
        programa_seleccionado=programa_filtro,
        usuario=usuario,
        es_admin=es_admin,
    )


# RUTA: vista individual del certificado de un estudiante
@app.route("/certificado/<identificacion>/<programa>")
@login_requerido
def pagina_certificado(identificacion, programa):
    usuario = obtener_usuario_actual()
    resultados = calcular_resultados()

    resultado = None  # VARIABLE: guardara el resultado buscado

    # BUCLE: busca el estudiante + programa pedidos en la URL
    for item in resultados:
        # CONDICIONAL: coincide identificacion y programa
        if item["identificacion"] == identificacion and item["programa"] == programa:
            resultado = item

    # CONDICIONAL: si no existe, se responde con error 404
    if resultado is None:
        abort(404)

    # CONDICIONAL: sin certificado no se muestra la vista de certificado
    if resultado["estado"] == "Sin certificado":
        abort(404)

    # VARIABLE: fecha de emision de hoy (se muestra en el certificado)
    from datetime import date
    fecha_emision = date.today().strftime("%d/%m/%Y")

    return render_template(
        "certificado.html",
        resultado=resultado,
        formatear=formatear_numero,
        fecha_emision=fecha_emision,
        usuario=usuario,
    )


# Punto de entrada: arranca la aplicacion si se ejecuta este archivo
if __name__ == "__main__":
    # debug=True muestra los errores en pantalla mientras se desarrolla
    app.run(debug=True)
